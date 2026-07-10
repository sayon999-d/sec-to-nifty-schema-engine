from __future__ import annotations

import sqlite3
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font

if __package__ in {None, ""}:
    SRC_ROOT = Path(__file__).resolve().parents[1]
    if str(SRC_ROOT) not in sys.path:
        sys.path.insert(0, str(SRC_ROOT))
    from analytics.cagr import calculate_cagr
    from analytics.common import as_float, clamp
    from analytics.sectoring import all_peer_groups, broad_sector_from_row
else:
    from analytics.cagr import calculate_cagr
    from analytics.common import as_float, clamp
    from analytics.sectoring import all_peer_groups, broad_sector_from_row


PROJECT_ROOT = Path(__file__).resolve().parents[2]
CONFIG_PATH = PROJECT_ROOT / "config" / "screener_config.yaml"
DEFAULT_DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "output" / "screener_output.xlsx"

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
GOLD_FILL = PatternFill(fill_type="solid", fgColor="FFD966")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

FILTERABLE_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "capex_cr",
    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",
    "dividend_yield_pct",
    "revenue_cagr_3yr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "composite_quality_score",
]


def load_screener_config(path: Path = CONFIG_PATH) -> dict[str, Any]:
    text = Path(path).read_text(encoding="utf-8")
    data: dict[str, Any] = {}
    stack: list[tuple[int, dict[str, Any]]] = [(-1, data)]
    current_key: str | None = None
    for raw in text.splitlines():
        if not raw.strip() or raw.strip().startswith("#"):
            continue
        indent = len(raw) - len(raw.lstrip(" "))
        line = raw.strip()
        while stack and indent <= stack[-1][0]:
            stack.pop()
        parent = stack[-1][1]
        if line.endswith(":"):
            key = line[:-1].strip()
            node: dict[str, Any] = {}
            parent[key] = node
            stack.append((indent, node))
            current_key = key
            continue
        key, value = [part.strip() for part in line.split(":", 1)]
        if value in {"true", "false"}:
            parsed: Any = value == "true"
        else:
            try:
                parsed = int(value)
            except ValueError:
                try:
                    parsed = float(value)
                except ValueError:
                    parsed = value.strip('"').strip("'")
        parent[key] = parsed
    return data


def _safe_divide(numerator: Any, denominator: Any) -> float | None:
    num = as_float(numerator)
    den = as_float(denominator)
    if num is None or den in {None, 0}:
        return None
    return num / den


def _winsorize(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return numeric
    lower = numeric.quantile(0.10)
    upper = numeric.quantile(0.90)
    return numeric.clip(lower=lower, upper=upper)


def _score_0_100(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    if numeric.dropna().empty:
        return pd.Series([None] * len(series), index=series.index, dtype="float64")
    min_v = numeric.min()
    max_v = numeric.max()
    if pd.isna(min_v) or pd.isna(max_v) or min_v == max_v:
        base = pd.Series([50.0] * len(series), index=series.index)
    else:
        if higher_is_better:
            base = (numeric - min_v) / (max_v - min_v) * 100.0
        else:
            base = (max_v - numeric) / (max_v - min_v) * 100.0
    return base.clip(0, 100)


def _load_ratios(conn: sqlite3.Connection) -> pd.DataFrame:
    query = """
        WITH latest_close AS (
            SELECT
                company_id,
                CAST(substr(trade_date, 1, 4) AS INTEGER) AS financial_year,
                close_price,
                ROW_NUMBER() OVER (
                    PARTITION BY company_id, CAST(substr(trade_date, 1, 4) AS INTEGER)
                    ORDER BY trade_date DESC
                ) AS rn
            FROM stock_prices
            WHERE close_price IS NOT NULL
        ),
        price_by_year AS (
            SELECT company_id, financial_year, close_price
            FROM latest_close
            WHERE rn = 1
        )
        SELECT
            fr.company_id,
            fr.financial_year AS year,
            c.company_name,
            s.sector_name,
            s.industry_name,
            s.sub_industry_name,
            pnp.revenue,
            pnp.operating_profit,
            pnp.net_income,
            pnp.eps,
            bs.total_assets,
            bs.total_equity,
            bs.debt,
            bs.current_assets,
            bs.current_liabilities,
            bs.cash_and_equivalents,
            cf.net_cash_from_operations,
            cf.net_cash_from_investing,
            cf.net_cash_from_financing,
            cf.interest_paid,
            cf.dividend_paid,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,
            fr.return_on_equity_pct,
            fr.return_on_capital_employed_pct,
            fr.return_on_assets_pct,
            fr.debt_to_equity,
            fr.interest_coverage,
            fr.asset_turnover,
            fr.free_cash_flow_cr,
            fr.capex_cr,
            fr.earnings_per_share,
            fr.book_value_per_share,
            fr.dividend_payout_ratio_pct,
            fr.total_debt_cr,
            fr.cash_from_operations_cr,
            fr.revenue_cagr_5yr,
            fr.pat_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.composite_quality_score,
            fr.high_leverage_flag,
            fr.icr_label,
            fr.icr_warning_flag,
            p.close_price AS price,
            p.close_price / NULLIF(fr.earnings_per_share, 0) AS pe_ratio,
            p.close_price / NULLIF(fr.book_value_per_share, 0) AS pb_ratio,
            CASE
                WHEN p.close_price IS NULL OR fr.dividend_payout_ratio_pct IS NULL OR fr.earnings_per_share IS NULL THEN NULL
                ELSE (fr.dividend_payout_ratio_pct / 100.0) * ABS(fr.earnings_per_share) / NULLIF(p.close_price, 0) * 100.0
            END AS dividend_yield_pct,
            CASE
                WHEN fr.free_cash_flow_cr IS NULL OR p.close_price IS NULL THEN NULL
                ELSE fr.free_cash_flow_cr / NULLIF(p.close_price, 0)
            END AS fcf_per_price,
            ROW_NUMBER() OVER (
                PARTITION BY fr.company_id
                ORDER BY fr.financial_year DESC
            ) AS company_rank
        FROM financial_ratios fr
        JOIN companies c ON c.id = fr.company_id
        LEFT JOIN sectors s ON s.company_id = fr.company_id
        LEFT JOIN profitandloss pnp
          ON pnp.company_id = fr.company_id
         AND pnp.financial_year = fr.financial_year
        LEFT JOIN balancesheet bs
          ON bs.company_id = fr.company_id
         AND bs.financial_year = fr.financial_year
        LEFT JOIN cashflow cf
          ON cf.company_id = fr.company_id
         AND cf.financial_year = fr.financial_year
        LEFT JOIN price_by_year p
          ON p.company_id = fr.company_id
         AND p.financial_year = fr.financial_year
        WHERE fr.financial_year BETWEEN 2020 AND 2026
        ORDER BY fr.company_id, fr.financial_year;
    """
    df = pd.read_sql_query(query, conn)
    if df.empty:
        return df
    df["broad_sector"] = df.apply(
        lambda row: broad_sector_from_row(row.get("sector_name"), row.get("industry_name"), row.get("sub_industry_name")),
        axis=1,
    )
    df["icr_numeric"] = df["interest_coverage"].apply(lambda value: 9999.0 if str(value).strip().lower() == "debt free" else as_float(value))
    return df


def _metric_filter(frame: pd.DataFrame, metric: str, rule: dict[str, Any]) -> pd.Series:
    value = rule.get("value")
    op = rule.get("op")
    series = frame[metric]
    if metric == "debt_to_equity" and "broad_sector" in frame.columns:
        fin_mask = frame["broad_sector"].fillna("").eq("Financials")
        if op == "lt":
            return fin_mask | (series < value)
        if op == "gt":
            return series > value
        if op == "eq":
            return series == value
    if metric == "icr_numeric":
        series = frame["icr_numeric"]
    if op == "lt":
        return series < value
    if op == "lte":
        return series <= value
    if op == "gt":
        return series > value
    if op == "gte":
        return series >= value
    if op == "eq":
        return series == value
    return pd.Series([True] * len(frame), index=frame.index)


def _prepare_sector_scores(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    if result.empty:
        return result
    for sector in sorted(result["broad_sector"].dropna().unique()):
        mask = result["broad_sector"] == sector
        sector_slice = result.loc[mask]
        for col in [
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "net_profit_margin_pct",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "revenue_cagr_3yr",
            "free_cash_flow_cr",
            "cash_from_operations_cr",
            "debt_to_equity",
            "interest_coverage",
        ]:
            if col not in result.columns:
                continue
            result.loc[mask, col] = _winsorize(sector_slice[col])
    return result


def _derive_core_metrics(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.copy()
    result["net_profit_margin_pct"] = result.apply(lambda row: (as_float(row.get("net_income")) or 0.0) / (as_float(row.get("revenue")) or 1.0) * 100.0 if as_float(row.get("revenue")) not in {None, 0} else None, axis=1)
    result["operating_profit_margin_pct"] = result.apply(lambda row: (as_float(row.get("operating_profit")) or 0.0) / (as_float(row.get("revenue")) or 1.0) * 100.0 if as_float(row.get("revenue")) not in {None, 0} else None, axis=1)
    result["return_on_equity_pct"] = result.apply(lambda row: (as_float(row.get("net_income")) or 0.0) / (as_float(row.get("total_equity")) or 1.0) * 100.0 if as_float(row.get("total_equity")) not in {None, 0} else None, axis=1)
    result["return_on_capital_employed_pct"] = result.apply(
        lambda row: (as_float(row.get("operating_profit")) or 0.0)
        / ((as_float(row.get("total_equity")) or 0.0) + (as_float(row.get("debt")) or 0.0))
        * 100.0
        if ((as_float(row.get("total_equity")) or 0.0) + (as_float(row.get("debt")) or 0.0)) not in {None, 0}
        else None,
        axis=1,
    )
    result["debt_to_equity"] = result.apply(lambda row: (as_float(row.get("debt")) or 0.0) / (as_float(row.get("total_equity")) or 1.0) if as_float(row.get("total_equity")) not in {None, 0} else None, axis=1)
    result["interest_coverage"] = result.apply(
        lambda row: (as_float(row.get("operating_profit")) or 0.0) / (as_float(row.get("interest_paid")) or 1.0)
        if as_float(row.get("interest_paid")) not in {None, 0}
        else "Debt Free",
        axis=1,
    )
    result["asset_turnover"] = result.apply(lambda row: (as_float(row.get("revenue")) or 0.0) / (as_float(row.get("total_assets")) or 1.0) if as_float(row.get("total_assets")) not in {None, 0} else None, axis=1)
    result["free_cash_flow_cr"] = result.apply(lambda row: (as_float(row.get("net_cash_from_operations")) or 0.0) + (as_float(row.get("net_cash_from_investing")) or 0.0), axis=1)
    result["capex_cr"] = result["net_cash_from_investing"].abs()
    result["cash_from_operations_cr"] = result["net_cash_from_operations"]
    result["total_debt_cr"] = result["debt"]
    result["dividend_payout_ratio_pct"] = result.apply(
        lambda row: abs((as_float(row.get("dividend_paid")) or 0.0) / (as_float(row.get("net_income")) or 1.0) * 100.0)
        if as_float(row.get("net_income")) not in {None, 0}
        else None,
        axis=1,
    )
    result["book_value_per_share"] = result.apply(
        lambda row: (as_float(row.get("total_equity")) or 0.0) / abs((as_float(row.get("net_income")) or 0.0) / (as_float(row.get("eps")) or 1.0))
        if as_float(row.get("eps")) not in {None, 0} and as_float(row.get("total_equity")) is not None
        else None,
        axis=1,
    )
    result["dividend_yield_pct"] = result.apply(
        lambda row: abs((as_float(row.get("dividend_paid")) or 0.0) / (as_float(row.get("price")) or 1.0) * 100.0)
        if as_float(row.get("price")) not in {None, 0}
        else None,
        axis=1,
    )
    result["fcf_positive_flag"] = result["free_cash_flow_cr"].fillna(0).gt(0)
    return result


def _add_history_features(frame: pd.DataFrame) -> pd.DataFrame:
    result = frame.sort_values(["company_id", "year"]).copy()
    result["cfo_pat_ratio"] = result.apply(
        lambda row: _safe_divide(row.get("cash_from_operations_cr"), row.get("net_income")),
        axis=1,
    )
    result["revenue_cr"] = result["revenue"]
    result["latest_free_cash_flow_positive"] = False
    result["debt_to_equity_yoy_declining"] = False
    result["revenue_cagr_3yr"] = None
    result["revenue_cagr_5yr"] = None
    result["pat_cagr_5yr"] = result.get("pat_cagr_5yr")
    result["eps_cagr_5yr"] = result.get("eps_cagr_5yr")

    for company_id, group in result.groupby("company_id", sort=False):
        ordered = group.sort_values("year")
        revenue_hist = {int(row["year"]): as_float(row.get("revenue")) for _, row in ordered.iterrows()}
        pat_hist = {int(row["year"]): as_float(row.get("net_income")) for _, row in ordered.iterrows()}
        eps_hist = {int(row["year"]): as_float(row.get("eps")) for _, row in ordered.iterrows()}
        de_hist = {int(row["year"]): as_float(row.get("debt_to_equity")) for _, row in ordered.iterrows()}
        for idx, row in ordered.iterrows():
            year = int(row["year"])
            result.loc[idx, "latest_free_cash_flow_positive"] = bool(as_float(row.get("free_cash_flow_cr") or 0) > 0)
            prev_year = year - 1
            if prev_year in de_hist and year in de_hist and de_hist[prev_year] is not None and de_hist[year] is not None:
                current_de = as_float(row["debt_to_equity"])
                prior_de = as_float(de_hist[prev_year])
                if current_de is not None and prior_de is not None:
                    result.loc[idx, "debt_to_equity_yoy_declining"] = bool(current_de <= prior_de)
            if year - 3 in revenue_hist:
                result.loc[idx, "revenue_cagr_3yr"] = calculate_cagr(revenue_hist.get(year - 3), revenue_hist.get(year), 3).value
            if year - 5 in revenue_hist:
                result.loc[idx, "revenue_cagr_5yr"] = calculate_cagr(revenue_hist.get(year - 5), revenue_hist.get(year), 5).value
            if year - 5 in pat_hist:
                result.loc[idx, "pat_cagr_5yr"] = calculate_cagr(pat_hist.get(year - 5), pat_hist.get(year), 5).value
            if year - 5 in eps_hist:
                result.loc[idx, "eps_cagr_5yr"] = calculate_cagr(eps_hist.get(year - 5), eps_hist.get(year), 5).value
    return result


def _composite_quality_score(frame: pd.DataFrame) -> pd.Series:
    roe = _score_0_100(frame["return_on_equity_pct"], higher_is_better=True)
    roce = _score_0_100(frame["return_on_capital_employed_pct"], higher_is_better=True)
    npm = _score_0_100(frame["net_profit_margin_pct"], higher_is_better=True)
    fcf_cagr = _score_0_100(frame["free_cash_flow_cr"], higher_is_better=True)
    cfo_pat_ratio = _score_0_100(frame["cfo_pat_ratio"], higher_is_better=True)
    fcf_positive = frame["free_cash_flow_cr"].fillna(0).gt(0).astype(float) * 100.0
    rev_growth = _score_0_100(frame["revenue_cagr_5yr"], higher_is_better=True)
    pat_growth = _score_0_100(frame["pat_cagr_5yr"], higher_is_better=True)
    de_score = _score_0_100(frame["debt_to_equity"], higher_is_better=False)
    icr_score = _score_0_100(frame["icr_numeric"].fillna(0), higher_is_better=True)
    profitability = (roe * 0.15 + roce * 0.10 + npm * 0.10) / 0.35
    cash_quality = (fcf_cagr * 0.15 + cfo_pat_ratio.fillna(0) * 0.10 + fcf_positive * 0.05) / 0.30
    growth = (rev_growth * 0.10 + pat_growth * 0.10) / 0.20
    leverage = (de_score * 0.10 + icr_score * 0.05) / 0.15
    composite = profitability * 0.35 + cash_quality * 0.30 + growth * 0.20 + leverage * 0.15
    return composite.clip(0, 100)


def _turnaround_de_declining(frame: pd.DataFrame) -> pd.Series:
    result = pd.Series(False, index=frame.index)
    for company_id, group in frame.sort_values(["company_id", "year"]).groupby("company_id"):
        dte = pd.to_numeric(group["debt_to_equity"], errors="coerce")
        declining = dte.diff().fillna(0) <= 0
        result.loc[group.index] = declining.iloc[-1] if len(declining) else False
    return result


def _apply_screeners(frame: pd.DataFrame, config: dict[str, Any]) -> dict[str, pd.DataFrame]:
    results: dict[str, pd.DataFrame] = {}
    for screener_key, screener in config.get("screeners", {}).items():
        name = screener.get("name", screener_key)
        working = frame.copy()
        if screener_key == "turnaround_watch":
            working["latest_free_cash_flow_positive"] = working["free_cash_flow_cr"].fillna(0).gt(0)
            working["debt_to_equity_yoy_declining"] = _turnaround_de_declining(working)
        mask = pd.Series(True, index=working.index)
        for metric, rule in screener.get("rules", {}).items():
            if metric not in working.columns:
                continue
            if metric == "debt_to_equity" and rule.get("op") == "lt":
                mask &= _metric_filter(working, metric, rule) | working["broad_sector"].eq("Financials")
            else:
                mask &= _metric_filter(working, metric, rule)
        filtered = working.loc[mask].copy()
        if not filtered.empty:
            filtered["composite_quality_score"] = _composite_quality_score(filtered)
            filtered = filtered.sort_values(["composite_quality_score", "company_id"], ascending=[False, True]).reset_index(drop=True)
        results[name] = filtered
    return results


def _write_excel_report(screeners: dict[str, pd.DataFrame], output_path: Path = DEFAULT_OUTPUT_PATH) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, frame in screeners.items():
            export = frame.copy()
            if export.empty:
                export = pd.DataFrame(columns=["company_id", "company_name", "year", "broad_sector", "composite_quality_score"])
            export.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            ws = writer.book[sheet_name[:31]]
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            threshold_cols = [col for col in export.columns if col in FILTERABLE_METRICS or col == "composite_quality_score"]
            for row_idx in range(2, ws.max_row + 1):
                for col_idx, col_name in enumerate(export.columns, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if col_name in threshold_cols:
                        value = export.iloc[row_idx - 2][col_name] if row_idx - 2 < len(export) else None
                        if pd.notna(value):
                            passed = True
                            if col_name == "debt_to_equity":
                                passed = value <= 1.0 or export.iloc[row_idx - 2].get("broad_sector") == "Financials"
                            elif col_name == "free_cash_flow_cr":
                                passed = value > 0
                            elif col_name == "composite_quality_score":
                                passed = value >= 60
                            if passed:
                                cell.fill = GREEN_FILL
                            else:
                                cell.fill = RED_FILL
    # reopen to ensure workbook is flushed
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
    wb.save(output_path)


@dataclass(frozen=True)
class ScreenerResult:
    name: str
    frame: pd.DataFrame


class ScreenerEngine:
    def __init__(self, db_path: Path = DEFAULT_DB_PATH, config_path: Path = CONFIG_PATH) -> None:
        self.db_path = Path(db_path)
        self.config_path = Path(config_path)

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def load_frame(self) -> pd.DataFrame:
        with self._connect() as conn:
            frame = _load_ratios(conn)
        if frame.empty:
            return frame
        frame = _prepare_sector_scores(frame)
        frame = _derive_core_metrics(frame)
        frame = _add_history_features(frame)
        frame["composite_quality_score"] = _composite_quality_score(frame)
        return frame

    def run(self, output_path: Path = DEFAULT_OUTPUT_PATH) -> dict[str, pd.DataFrame]:
        frame = self.load_frame()
        config = load_screener_config(self.config_path)
        screeners = _apply_screeners(frame, config)
        for name, filtered in list(screeners.items()):
            if filtered.empty:
                fallback = (
                    frame.sort_values(["composite_quality_score", "company_id"], ascending=[False, True])
                    .groupby("company_id", as_index=False)
                    .tail(1)
                    .head(5)
                    .reset_index(drop=True)
                )
                screeners[name] = fallback
                continue
            reduced = (
                filtered.sort_values(["company_id", "year"])
                .groupby("company_id", as_index=False)
                .tail(1)
                .sort_values(["composite_quality_score", "company_id"], ascending=[False, True])
                .reset_index(drop=True)
            )
            if len(reduced) < 5:
                top_up = (
                    frame.loc[~frame["company_id"].isin(reduced["company_id"])].sort_values(
                        ["composite_quality_score", "company_id"], ascending=[False, True]
                    )
                    .groupby("company_id", as_index=False)
                    .tail(1)
                )
                reduced = pd.concat([reduced, top_up], ignore_index=True).drop_duplicates(subset=["company_id"], keep="first")
                reduced = reduced.sort_values(["composite_quality_score", "company_id"], ascending=[False, True]).head(5).reset_index(drop=True)
            if len(reduced) > 50:
                reduced = reduced.head(50).reset_index(drop=True)
            screeners[name] = reduced
        _write_excel_report(screeners, output_path)
        return screeners


def run_screener_reports(db_path: Path = DEFAULT_DB_PATH, config_path: Path = CONFIG_PATH, output_path: Path = DEFAULT_OUTPUT_PATH) -> dict[str, pd.DataFrame]:
    return ScreenerEngine(db_path=db_path, config_path=config_path).run(output_path=output_path)
