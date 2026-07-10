from __future__ import annotations

import math
import sqlite3
import struct
import sys
import zlib
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

if __package__ in {None, ""}:
    SRC_ROOT = Path(__file__).resolve().parents[1]
    if str(SRC_ROOT) not in sys.path:
        sys.path.insert(0, str(SRC_ROOT))
    from analytics.common import as_float, ensure_parent
    from analytics.sectoring import all_peer_groups, peer_group_from_row
else:
    from .common import as_float, ensure_parent
    from .sectoring import all_peer_groups, peer_group_from_row


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"
DEFAULT_OUTPUT_PATH = PROJECT_ROOT / "output" / "peer_comparison.xlsx"
DEFAULT_RADAR_DIR = PROJECT_ROOT / "reports" / "radar_charts"
DEFAULT_LOG_PATH = PROJECT_ROOT / "output" / "peer_percentiles.log"

GREEN_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFEB9C")
RED_FILL = PatternFill(fill_type="solid", fgColor="FFC7CE")
GOLD_FILL = PatternFill(fill_type="solid", fgColor="FFD966")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)

PEER_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "composite_quality_score",
]

RADAR_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "debt_to_equity",
    "interest_coverage",
    "composite_quality_score",
]


def _log(message: str, path: Path = DEFAULT_LOG_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("a", encoding="utf-8") as handle:
        handle.write(message.rstrip() + "\n")


def _connect(db_path: Path = DEFAULT_DB_PATH) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    return conn


def _latest_ratios(conn: sqlite3.Connection) -> pd.DataFrame:
    query = """
        WITH latest AS (
            SELECT company_id, MAX(financial_year) AS financial_year
            FROM financial_ratios
            WHERE financial_year BETWEEN 2020 AND 2026
            GROUP BY company_id
        )
        SELECT
            fr.company_id,
            c.company_name,
            fr.financial_year AS year,
            pnp.revenue,
            pnp.operating_profit,
            pnp.net_income,
            pnp.eps,
            bs.total_assets,
            bs.total_equity,
            bs.debt,
            cf.net_cash_from_operations,
            cf.net_cash_from_investing,
            cf.interest_paid,
            cf.dividend_paid,
            fr.return_on_equity_pct,
            fr.return_on_capital_employed_pct,
            fr.net_profit_margin_pct,
            fr.operating_profit_margin_pct,
            fr.revenue_cagr_5yr,
            fr.pat_cagr_5yr,
            fr.eps_cagr_5yr,
            fr.debt_to_equity,
            fr.interest_coverage,
            fr.composite_quality_score,
            fr.free_cash_flow_cr,
            fr.cash_from_operations_cr,
            fr.capex_cr,
            fr.earnings_per_share,
            fr.book_value_per_share,
            fr.dividend_payout_ratio_pct,
            fr.total_debt_cr,
            fr.high_leverage_flag,
            fr.icr_label,
            fr.icr_warning_flag,
            s.sector_name,
            s.industry_name,
            s.sub_industry_name
        FROM financial_ratios fr
        JOIN latest l ON l.company_id = fr.company_id AND l.financial_year = fr.financial_year
        JOIN companies c ON c.id = fr.company_id
        LEFT JOIN sectors s ON s.company_id = fr.company_id
        LEFT JOIN profitandloss pnp
          ON pnp.company_id = fr.company_id
         AND pnp.financial_year = fr.financial_year
        LEFT JOIN balancesheet bs
          ON bs.company_id = fr.company_id
         AND bs.financial_year = fr.financial_year
        LEFT JOIN cashflow cf
          ON cf.company_id = fr.company_id
         AND cf.financial_year = fr.financial_year
        ORDER BY fr.company_id;
    """
    df = pd.read_sql_query(query, conn)
    if df.empty:
        return df
    df["peer_group_name"] = df.apply(lambda row: peer_group_from_row(row.get("sector_name"), row.get("industry_name"), row.get("sub_industry_name")), axis=1)
    df["icr_numeric"] = df["interest_coverage"].apply(lambda value: 9999.0 if str(value).strip().lower() == "debt free" else as_float(value))
    df["net_profit_margin_pct"] = df.apply(lambda row: (as_float(row.get("net_income")) or 0.0) / (as_float(row.get("revenue")) or 1.0) * 100.0 if as_float(row.get("revenue")) not in {None, 0} else None, axis=1)
    df["operating_profit_margin_pct"] = df.apply(lambda row: (as_float(row.get("operating_profit")) or 0.0) / (as_float(row.get("revenue")) or 1.0) * 100.0 if as_float(row.get("revenue")) not in {None, 0} else None, axis=1)
    df["return_on_equity_pct"] = df.apply(lambda row: (as_float(row.get("net_income")) or 0.0) / (as_float(row.get("total_equity")) or 1.0) * 100.0 if as_float(row.get("total_equity")) not in {None, 0} else None, axis=1)
    df["return_on_capital_employed_pct"] = df.apply(
        lambda row: (as_float(row.get("operating_profit")) or 0.0)
        / ((as_float(row.get("total_equity")) or 0.0) + (as_float(row.get("debt")) or 0.0))
        * 100.0
        if ((as_float(row.get("total_equity")) or 0.0) + (as_float(row.get("debt")) or 0.0)) not in {None, 0}
        else None,
        axis=1,
    )
    df["debt_to_equity"] = df.apply(lambda row: (as_float(row.get("debt")) or 0.0) / (as_float(row.get("total_equity")) or 1.0) if as_float(row.get("total_equity")) not in {None, 0} else None, axis=1)
    df["free_cash_flow_cr"] = df.apply(lambda row: (as_float(row.get("net_cash_from_operations")) or 0.0) + (as_float(row.get("net_cash_from_investing")) or 0.0), axis=1)
    df["composite_quality_score"] = df.apply(
        lambda row: min(
            100.0,
            max(
                0.0,
                sum(
                    [
                        25.0 if (as_float(row.get("return_on_equity_pct")) or 0.0) > 0 else 0.0,
                        25.0 if (as_float(row.get("return_on_capital_employed_pct")) or 0.0) > 0 else 0.0,
                        25.0 if (as_float(row.get("net_profit_margin_pct")) or 0.0) > 0 else 0.0,
                        25.0 if (as_float(row.get("free_cash_flow_cr")) or 0.0) > 0 else 0.0,
                    ]
                ),
            ),
        ),
        axis=1,
    )
    return df


def _safe_percent_rank(series: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(series, errors="coerce")
    non_null = numeric.dropna()
    if non_null.empty:
        return pd.Series([None] * len(series), index=series.index, dtype="float64")
    if len(non_null) == 1:
        result = pd.Series([1.0 if pd.notna(v) else None for v in numeric], index=series.index, dtype="float64")
        return result
    ranks = numeric.rank(method="min", pct=False)
    return ((ranks - 1) / (len(non_null) - 1)).where(numeric.notna())


def _metric_percentile(frame: pd.DataFrame, metric: str) -> pd.Series:
    series = pd.to_numeric(frame[metric], errors="coerce")
    if metric == "debt_to_equity":
        series = -series
    return _safe_percent_rank(series)


def _ensure_peer_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS peer_percentiles (
            company_id INTEGER NOT NULL,
            peer_group_name TEXT NOT NULL,
            metric TEXT NOT NULL,
            value REAL,
            percentile_rank REAL,
            year INTEGER NOT NULL
        );
        """
    )
    conn.execute("DELETE FROM peer_percentiles;")
    conn.commit()


def build_peer_percentiles(db_path: Path = DEFAULT_DB_PATH) -> pd.DataFrame:
    with _connect(db_path) as conn:
        frame = _latest_ratios(conn)
        _ensure_peer_table(conn)
        rows: list[dict[str, Any]] = []
        if frame.empty:
            return pd.DataFrame()
        for group_name, group in frame.groupby("peer_group_name", dropna=False):
            if not isinstance(group_name, str) or not group_name.strip() or group_name == "Unknown":
                _log("No peer group assigned")
                continue
            for metric in PEER_METRICS:
                if metric not in group.columns:
                    continue
                ranks = _metric_percentile(group, metric)
                for idx, (_, row) in enumerate(group.iterrows()):
                    rows.append(
                        {
                            "company_id": int(row["company_id"]),
                            "peer_group_name": group_name,
                            "metric": metric,
                            "value": None if pd.isna(row[metric]) else float(as_float(row[metric]) or 0.0),
                            "percentile_rank": None if pd.isna(ranks.iloc[idx]) else float(ranks.iloc[idx]),
                            "year": int(row["year"]),
                        }
                    )
        if rows:
            export = pd.DataFrame(rows)
            export.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        else:
            export = pd.DataFrame(columns=["company_id", "peer_group_name", "metric", "value", "percentile_rank", "year"])
        conn.commit()
        return export


def _sheet_data(frame: pd.DataFrame, group_name: str) -> pd.DataFrame:
    subset = frame.loc[frame["peer_group_name"] == group_name].copy()
    if subset.empty:
        return subset
    percentile_cols = {}
    for metric in PEER_METRICS:
        ranks = _metric_percentile(subset, metric)
        percentile_cols[f"{metric}_percentile"] = ranks
    for key, value in percentile_cols.items():
        subset[key] = value
    subset["benchmark_flag"] = subset["return_on_equity_pct_percentile"].eq(subset["return_on_equity_pct_percentile"].max())
    subset = subset.sort_values(["benchmark_flag", "composite_quality_score"], ascending=[False, False]).reset_index(drop=True)
    return subset


def _write_peer_excel(frame: pd.DataFrame, output_path: Path = DEFAULT_OUTPUT_PATH) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    all_groups = all_peer_groups()
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for group_name in all_groups:
            subset = _sheet_data(frame, group_name)
            if subset.empty:
                subset = pd.DataFrame(columns=["company_id", "company_name", "year"] + PEER_METRICS + [f"{m}_percentile" for m in PEER_METRICS])
            export_cols = ["company_id", "company_name", "year"] + PEER_METRICS + [f"{m}_percentile" for m in PEER_METRICS]
            subset = subset.reindex(columns=export_cols)
            median_row = {col: None for col in export_cols}
            median_row["company_name"] = "Group Median"
            median_row["company_id"] = None
            median_row["year"] = None
            for col in PEER_METRICS:
                median_row[col] = pd.to_numeric(subset[col], errors="coerce").median() if col in subset else None
                median_row[f"{col}_percentile"] = pd.to_numeric(subset[f"{col}_percentile"], errors="coerce").median() if f"{col}_percentile" in subset else None
            subset = pd.concat([subset, pd.DataFrame([median_row])], ignore_index=True)
            subset.to_excel(writer, sheet_name=group_name[:31], index=False)
            ws = writer.book[group_name[:31]]
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
            for row_idx in range(2, ws.max_row + 1):
                row_company = ws.cell(row=row_idx, column=2).value
                if row_company == "Group Median":
                    for cell in ws[row_idx]:
                        cell.fill = GOLD_FILL
                    continue
                if row_idx == 2:
                    for cell in ws[row_idx]:
                        cell.fill = GOLD_FILL
                for col_idx, col_name in enumerate(ws[1], start=1):
                    header = col_name.value
                    if header and header.endswith("_percentile"):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        if value is None:
                            continue
                        try:
                            numeric_value = float(value)
                        except Exception:
                            continue
                        if numeric_value >= 0.75:
                            cell.fill = GREEN_FILL
                        elif numeric_value <= 0.25:
                            cell.fill = RED_FILL
                        else:
                            cell.fill = YELLOW_FILL
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
    wb.save(output_path)


def _normalize(values: list[float | None]) -> list[float]:
    series = pd.Series(values, dtype="float64")
    if series.dropna().empty:
        return [0.5 for _ in values]
    lower = series.quantile(0.10)
    upper = series.quantile(0.90)
    clipped = series.clip(lower=lower, upper=upper)
    if lower == upper:
        return [0.5 if pd.notna(v) else 0.0 for v in clipped]
    return [float((v - lower) / (upper - lower)) if pd.notna(v) else 0.0 for v in clipped]


def _write_png(path: Path, width: int, height: int, pixels: list[tuple[int, int, int, int]]) -> None:
    raw = bytearray()
    for y in range(height):
        raw.append(0)
        for x in range(width):
            raw.extend(bytes(pixels[y * width + x]))
    compressor = zlib.compress(bytes(raw))

    def chunk(tag: bytes, data: bytes) -> bytes:
        return struct.pack("!I", len(data)) + tag + data + struct.pack("!I", zlib.crc32(tag + data) & 0xFFFFFFFF)

    png = b"\x89PNG\r\n\x1a\n"
    png += chunk(b"IHDR", struct.pack("!IIBBBBB", width, height, 8, 6, 0, 0, 0))
    png += chunk(b"IDAT", compressor)
    png += chunk(b"IEND", b"")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(png)


def _draw_line(pixels: list[tuple[int, int, int, int]], width: int, height: int, x0: int, y0: int, x1: int, y1: int, color: tuple[int, int, int, int]) -> None:
    dx = abs(x1 - x0)
    dy = -abs(y1 - y0)
    sx = 1 if x0 < x1 else -1
    sy = 1 if y0 < y1 else -1
    err = dx + dy
    while True:
        if 0 <= x0 < width and 0 <= y0 < height:
            pixels[y0 * width + x0] = color
        if x0 == x1 and y0 == y1:
            break
        e2 = 2 * err
        if e2 >= dy:
            err += dy
            x0 += sx
        if e2 <= dx:
            err += dx
            y0 += sy


def _fill_polygon(pixels: list[tuple[int, int, int, int]], width: int, height: int, points: list[tuple[int, int]], color: tuple[int, int, int, int]) -> None:
    if not points:
        return
    min_y = max(min(y for _, y in points), 0)
    max_y = min(max(y for _, y in points), height - 1)
    for y in range(min_y, max_y + 1):
        intersections: list[int] = []
        for i in range(len(points)):
            x1, y1 = points[i]
            x2, y2 = points[(i + 1) % len(points)]
            if y1 == y2:
                continue
            if y < min(y1, y2) or y >= max(y1, y2):
                continue
            x = int(x1 + (y - y1) * (x2 - x1) / (y2 - y1))
            intersections.append(x)
        intersections.sort()
        for i in range(0, len(intersections), 2):
            if i + 1 >= len(intersections):
                break
            for x in range(intersections[i], intersections[i + 1] + 1):
                if 0 <= x < width:
                    pixels[y * width + x] = color


def _radar_chart(company_name: str, values: list[float], average: list[float], output_path: Path) -> None:
    width, height = 900, 900
    pixels = [(255, 255, 255, 255)] * (width * height)
    center = (width // 2, height // 2)
    radius = 300
    axes = len(values)
    angles = [2 * math.pi * i / axes - math.pi / 2 for i in range(axes)]

    for i, angle in enumerate(angles):
        x = int(center[0] + radius * math.cos(angle))
        y = int(center[1] + radius * math.sin(angle))
        _draw_line(pixels, width, height, center[0], center[1], x, y, (200, 200, 200, 255))

    def points_for(dataset: list[float], scale: float = 1.0) -> list[tuple[int, int]]:
        pts = []
        for value, angle in zip(dataset, angles):
            r = radius * max(0.0, min(1.0, value)) * scale
            x = int(center[0] + r * math.cos(angle))
            y = int(center[1] + r * math.sin(angle))
            pts.append((x, y))
        return pts

    company_pts = points_for(values)
    avg_pts = points_for(average)
    _fill_polygon(pixels, width, height, company_pts, (198, 239, 206, 120))
    _fill_polygon(pixels, width, height, avg_pts, (255, 255, 255, 0))
    for i in range(len(company_pts)):
        _draw_line(pixels, width, height, *company_pts[i], *company_pts[(i + 1) % len(company_pts)], (34, 139, 34, 255))
        _draw_line(pixels, width, height, *avg_pts[i], *avg_pts[(i + 1) % len(avg_pts)], (0, 0, 139, 255))
    _write_png(output_path, width, height, pixels)


def generate_peer_reports(db_path: Path = DEFAULT_DB_PATH, output_path: Path = DEFAULT_OUTPUT_PATH, radar_dir: Path = DEFAULT_RADAR_DIR) -> pd.DataFrame:
    with _connect(db_path) as conn:
        frame = _latest_ratios(conn)
        if frame.empty:
            return frame
        _ensure_peer_table(conn)
        rows: list[dict[str, Any]] = []
        for group_name, group in frame.groupby("peer_group_name", dropna=False):
            if not isinstance(group_name, str) or not group_name.strip() or group_name == "Unknown":
                _log("No peer group assigned")
                continue
            for metric in PEER_METRICS:
                ranks = _metric_percentile(group, metric)
                for idx, (_, row) in enumerate(group.iterrows()):
                    rows.append(
                        {
                            "company_id": int(row["company_id"]),
                            "peer_group_name": group_name,
                            "metric": metric,
                            "value": None if pd.isna(row[metric]) else float(as_float(row[metric]) or 0.0),
                            "percentile_rank": None if pd.isna(ranks.iloc[idx]) else float(ranks.iloc[idx]),
                            "year": int(row["year"]),
                        }
                    )
        if rows:
            pd.DataFrame(rows).to_sql("peer_percentiles", conn, if_exists="append", index=False)
        _write_peer_excel(frame, output_path)

        radar_dir.mkdir(parents=True, exist_ok=True)
        group_means = {group: frame.loc[frame["peer_group_name"] == group, RADAR_METRICS].apply(pd.to_numeric, errors="coerce").mean().tolist() for group in frame["peer_group_name"].dropna().unique()}
        nifty_reference = frame[RADAR_METRICS].apply(pd.to_numeric, errors="coerce").mean().tolist()
        for _, row in frame.iterrows():
            group_name = row["peer_group_name"]
            reference = group_means.get(group_name, nifty_reference)
            if any(pd.isna(v) for v in reference):
                reference = nifty_reference
            values = [as_float(row.get(metric)) for metric in RADAR_METRICS]
            normalized = _normalize(values)
            reference_norm = _normalize(reference)
            output_file = radar_dir / f"{int(row['company_id'])}_radar.png"
            _radar_chart(str(row["company_name"]), normalized, reference_norm, output_file)
        conn.commit()
        return pd.DataFrame(rows)
